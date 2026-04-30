#!/usr/bin/env python3
"""
2026春季 鲜奶+常温奶 业务管理表 生成脚本
输出: 2026春季学生奶业务管理总表.xlsx
"""

import os, glob, re, warnings, traceback
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import defaultdict
import datetime

warnings.filterwarnings('ignore')

BASE_DIR = "/mnt/d/claude-code-project/data-treat"

# ─── Styles ───────────────────────────────────────────────────────────────────
def mk_fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def mk_font(bold=False, color="000000", size=10): return Font(bold=bold, color=color, size=size)

FILL = {
    "navy":   mk_fill("1F4E79"),
    "blue":   mk_fill("2E75B6"),
    "lblue":  mk_fill("D6E4F0"),
    "dblue":  mk_fill("BDD7EE"),
    "green":  mk_fill("375623"),
    "lgreen": mk_fill("E2EFDA"),
    "orange": mk_fill("C65911"),
    "lorange":mk_fill("FCE4D6"),
    "gray":   mk_fill("D9D9D9"),
    "lgray":  mk_fill("F2F2F2"),
    "white":  mk_fill("FFFFFF"),
    "yellow": mk_fill("FFF2CC"),
    "red":    mk_fill("C00000"),
    "lred":   mk_fill("FFE7E7"),
    "purple": mk_fill("7030A0"),
}
FONT = {
    "title":    mk_font(bold=True, color="FFFFFF", size=13),
    "hdr":      mk_font(bold=True, color="FFFFFF", size=10),
    "hdr_dark": mk_font(bold=True, color="FFFFFF", size=10),
    "sub_hdr":  mk_font(bold=True, color="1F4E79", size=10),
    "route":    mk_font(bold=True, color="FFFFFF", size=10),
    "bold":     mk_font(bold=True, size=10),
    "normal":   mk_font(size=10),
    "small":    mk_font(size=9),
    "total":    mk_font(bold=True, size=10),
}
thin = Side(style="thin",   color="999999")
thick = Side(style="medium", color="555555")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
THICK_BORDER = Border(left=thick, right=thick, top=thick, bottom=thick)
ALIGN = {
    "c": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "l": Alignment(horizontal="left",   vertical="center", wrap_text=True),
    "r": Alignment(horizontal="right",  vertical="center"),
}

def style(cell, fill=None, font=None, align="c", border=True):
    if fill:  cell.fill = fill
    if font:  cell.font = font
    cell.alignment = ALIGN.get(align, ALIGN["c"])
    if border: cell.border = BORDER

def header_row(ws, row, col_start, col_end, fill, font="hdr"):
    for c in range(col_start, col_end + 1):
        style(ws.cell(row=row, column=c), FILL[fill], FONT[font])

def data_row(ws, row, col_start, col_end, fill=None, font="normal"):
    for c in range(col_start, col_end + 1):
        style(ws.cell(row=row, column=c), FILL.get(fill), FONT[font])

def write(ws, row, col, val, fill=None, font=None, align="c", border=True, numfmt=None):
    cell = ws.cell(row=row, column=col, value=val)
    style(cell, FILL.get(fill) if fill else None,
          FONT.get(font) if font else None, align, border)
    if numfmt: cell.number_format = numfmt
    return cell

def merge_write(ws, r1, c1, r2, c2, val, fill=None, font=None, align="c"):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=val)
    style(cell, FILL.get(fill) if fill else None,
          FONT.get(font) if font else None, align)
    return cell

def set_col_widths(ws, widths):
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

def set_row_height(ws, row, h):
    ws.row_dimensions[row].height = h

# ─── Helper: Excel serial → date ─────────────────────────────────────────────
def to_date(v):
    if isinstance(v, (int, float)) and 40000 < v < 55000:
        return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    return None

# ─── Route config ─────────────────────────────────────────────────────────────
ROUTES = {
    "东乡线":   os.path.join(BASE_DIR, "2026春鲜奶东乡线"),
    "市区+英红": os.path.join(BASE_DIR, "2026春鲜奶市区+英红"),
    "西乡线":   os.path.join(BASE_DIR, "2026春鲜奶西乡线"),
    "黎-沙线":  os.path.join(BASE_DIR, "2026春鲜奶黎-沙线"),
}
ROUTE_ORDER = ["东乡线", "市区+英红", "西乡线", "黎-沙线"]
ROUTE_FILLS = {
    "东乡线":   ("navy",  "lblue"),
    "市区+英红": ("blue",  "dblue"),
    "西乡线":   ("green", "lgreen"),
    "黎-沙线":  ("orange","lorange"),
}

# ═══════════════════════════════════════════════════════════════════════════════
# PART 1 — 鲜奶 data from order import files (3天/4天)
# ═══════════════════════════════════════════════════════════════════════════════

def read_order_import(path, label):
    """Read the ERP order import file, return per-school per-product totals."""
    try:
        df = pd.read_excel(path, header=1, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        # Find key columns
        col_customer = next((c for c in df.columns if '客户' in c), None)
        col_product  = next((c for c in df.columns if '商品' in c or '名称' in c and '客户' not in c), None)
        col_qty      = next((c for c in df.columns if '数量' in c), None)
        col_note     = next((c for c in df.columns if '备注' in c and '整单' not in c), None)

        if not (col_customer and col_product and col_qty):
            print(f"  [WARN] {label}: can't find key columns. Cols: {list(df.columns)}")
            return pd.DataFrame()

        df = df[[col_customer, col_product, col_qty]].copy()
        df.columns = ['学校', '商品', '数量']
        df = df.dropna(subset=['学校'])
        df['数量'] = pd.to_numeric(df['数量'], errors='coerce').fillna(0)
        df['类型'] = df['商品'].apply(
            lambda x: '乳酸' if any(k in str(x) for k in ['乳酸', '乳饮', '酸奶']) else '鲜奶'
        )
        return df.groupby(['学校', '类型'])['数量'].sum().reset_index()
    except Exception as e:
        print(f"  [ERR] {label}: {e}")
        return pd.DataFrame()

path_3d = os.path.join(BASE_DIR, "3天鲜奶量录入明细总表.xlsx")
path_4d = os.path.join(BASE_DIR, "4天鲜奶量录入明细总表.xlsx")

raw_3d = read_order_import(path_3d, "3天订单")
raw_4d = read_order_import(path_4d, "4天订单")

def pivot_order(df, day):
    if df.empty: return pd.DataFrame(columns=['学校', f'{day}_鲜', f'{day}_酸'])
    p = df.pivot_table(index='学校', columns='类型', values='数量', aggfunc='sum').reset_index()
    p.columns.name = None
    p = p.rename(columns={'鲜奶': f'{day}_鲜', '乳酸': f'{day}_酸'})
    for col in [f'{day}_鲜', f'{day}_酸']:
        if col not in p.columns: p[col] = 0
    p[[f'{day}_鲜', f'{day}_酸']] = p[[f'{day}_鲜', f'{day}_酸']].fillna(0).astype(int)
    p[f'{day}_合计'] = p[f'{day}_鲜'] + p[f'{day}_酸']
    return p

pivot_3d = pivot_order(raw_3d, '周一')
pivot_4d = pivot_order(raw_4d, '周四')

# Merge into one delivery summary per school
if pivot_3d.empty and pivot_4d.empty:
    delivery_df = pd.DataFrame(columns=['学校','周一_鲜','周一_酸','周一_合计','周四_鲜','周四_酸','周四_合计','周合计'])
elif pivot_3d.empty:
    delivery_df = pivot_4d.copy()
    for c in ['周一_鲜','周一_酸','周一_合计']: delivery_df[c] = 0
elif pivot_4d.empty:
    delivery_df = pivot_3d.copy()
    for c in ['周四_鲜','周四_酸','周四_合计']: delivery_df[c] = 0
else:
    delivery_df = pd.merge(pivot_3d, pivot_4d, on='学校', how='outer').fillna(0)
    for c in ['周一_鲜','周一_酸','周一_合计','周四_鲜','周四_酸','周四_合计']:
        delivery_df[c] = delivery_df[c].astype(int)

delivery_df['周合计'] = delivery_df['周一_合计'] + delivery_df['周四_合计']

print(f"[订单] 3天: {len(pivot_3d)} 学校, 4天: {len(pivot_4d)} 学校, 合并: {len(delivery_df)} 学校")

# ═══════════════════════════════════════════════════════════════════════════════
# PART 2 — 鲜奶 subscription stats from individual school files
# ═══════════════════════════════════════════════════════════════════════════════

SUB_SHEET_NAMES = ['接龙表', '接龙', '老师', '对接', '订购']

def find_sub_sheet(wb):
    for name in SUB_SHEET_NAMES:
        if name in wb.sheetnames: return wb[name]
    for sn in wb.sheetnames:
        if any(k in sn for k in SUB_SHEET_NAMES): return wb[sn]
    return None

def parse_subscription(ws):
    """Extract subscription counts from a school's subscription sheet."""
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows: return None

    # Find the data start row (after headers, look for seq+name row)
    data_start = 5  # default
    for i, row in enumerate(all_rows[:10]):
        row_str = ' '.join(str(v) for v in row if v is not None)
        # Header rows contain '姓名' but not data
        if ('姓名' in row_str or '职工' in row_str) and ('序号' in row_str or '电话' in row_str):
            data_start = i + 2  # skip 1-2 header lines
            break

    # Detect which columns hold 鲜 and 酸 checkboxes by scanning header area
    # Heuristic: in most files col D=鲜, col E=酸 (0-indexed: 3, 4)
    fresh_col = 3
    lactic_col = 4

    # Scan header rows to confirm
    for row in all_rows[:data_start]:
        for ci, val in enumerate(row):
            v = str(val).strip() if val else ''
            if '鲜奶' in v and ci > 2: fresh_col = ci; break
        for ci, val in enumerate(row):
            v = str(val).strip() if val else ''
            if '乳酸' in v and ci > 2: lactic_col = ci; break

    CHECK_VALS = {'√', '✓', 'v', 'V', '1', '是', 'Y', 'y', '✔'}

    def is_checked(v):
        if v is None: return False
        s = str(v).strip()
        return s in CHECK_VALS or (len(s) > 0 and s[0] in CHECK_VALS)

    total = fresh_only = lactic_only = mixed = 0
    errors = 0

    for row in all_rows[data_start:]:
        if not row or len(row) < 2: continue
        name = row[1] if len(row) > 1 else None
        if name is None: continue
        name_str = str(name).strip()
        if not name_str or name_str in ('None', '', '合计', '小计', '总计', '合 计', '备注'):
            continue
        if any(name_str.startswith(k) for k in ('共', '注', '说明', '小结', '—', '-', '合计')):
            continue
        if not re.search(r'[一-鿿]', name_str):  # must have Chinese chars
            continue

        f_val = row[fresh_col]  if len(row) > fresh_col  else None
        a_val = row[lactic_col] if len(row) > lactic_col else None
        has_fresh  = is_checked(f_val)
        has_lactic = is_checked(a_val)

        if not has_fresh and not has_lactic:
            # Fallback: if person has name + phone, assume they subscribed
            # Some files only have name+quantity without checkbox
            errors += 1
            continue

        total += 1
        if has_fresh and has_lactic: mixed += 1
        elif has_fresh:              fresh_only += 1
        else:                        lactic_only += 1

    # If too many errors, try without checkbox detection (use any non-empty in data cols)
    if total == 0 and errors > 0:
        total = errors
        fresh_only = errors

    # Count fresh boxes (1份=100支: mixed→50鲜+50酸, fresh_only→100鲜, lactic_only→100酸)
    boxes_fresh  = (fresh_only * 100) + (mixed * 50)
    boxes_lactic = (lactic_only * 100) + (mixed * 50)
    total_subs   = fresh_only + lactic_only + mixed

    return {
        'people':      total,
        'fresh_only':  fresh_only,
        'lactic_only': lactic_only,
        'mixed':       mixed,
        'total_subs':  total_subs,
        'boxes_fresh':  boxes_fresh,
        'boxes_lactic': boxes_lactic,
        'boxes_total':  boxes_fresh + boxes_lactic,
        # Weekly per-delivery: 3天→3盒/份, 4天→4盒/份
        'mon_fresh':  round(boxes_fresh  / 100 * 3),
        'mon_lactic': round(boxes_lactic / 100 * 3),
        'thu_fresh':  round(boxes_fresh  / 100 * 4),
        'thu_lactic': round(boxes_lactic / 100 * 4),
    }

def clean_school_name(filename):
    name = os.path.basename(filename)
    name = re.sub(r'2026春季?', '', name)
    name = re.sub(r'（.+）', '', name)
    name = re.sub(r'\(.+\)', '', name)
    name = re.sub(r'教职工.+', '', name)
    name = re.sub(r'职工.+', '', name)
    name = re.sub(r'教师.+', '', name)
    name = re.sub(r'\.xlsx$', '', name, flags=re.I)
    name = re.sub(r'\.\d+\.xlsx$', '', name, flags=re.I)
    return name.strip()

SKIP_PATTERNS = ['备用', '量少', '空白', '~$', '.tmp']

school_records = []
for route_name in ROUTE_ORDER:
    route_dir = ROUTES[route_name]
    files = sorted(glob.glob(os.path.join(route_dir, "*.xlsx")))
    for fp in files:
        fname = os.path.basename(fp)
        if any(p in fname for p in SKIP_PATTERNS): continue
        school = clean_school_name(fp)
        rec = {'路线': route_name, '学校': school, '文件': fname}
        try:
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            sh = find_sub_sheet(wb)
            info = parse_subscription(sh) if sh else None
            wb.close()
            if info:
                rec.update(info)
            else:
                rec.update({'people':0,'fresh_only':0,'lactic_only':0,'mixed':0,
                            'total_subs':0,'boxes_fresh':0,'boxes_lactic':0,'boxes_total':0,
                            'mon_fresh':0,'mon_lactic':0,'thu_fresh':0,'thu_lactic':0})
        except Exception as e:
            print(f"  [ERR] {fname}: {e}")
            rec.update({'people':0,'fresh_only':0,'lactic_only':0,'mixed':0,
                        'total_subs':0,'boxes_fresh':0,'boxes_lactic':0,'boxes_total':0,
                        'mon_fresh':0,'mon_lactic':0,'thu_fresh':0,'thu_lactic':0})
        school_records.append(rec)

school_df = pd.DataFrame(school_records)
print(f"[鲜奶] 解析学校文件: {len(school_df)} 个")

# ═══════════════════════════════════════════════════════════════════════════════
# PART 3 — 常温奶 (student milk) from master file
# ═══════════════════════════════════════════════════════════════════════════════

master_path = os.path.join(BASE_DIR, "2026年春季各学校明细配送表.xlsx")

def parse_student_milk(path):
    """Parse 学生奶 sheet from master file."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if '学生奶' not in wb.sheetnames:
            wb.close()
            return pd.DataFrame()
        ws = wb['学生奶']

        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            wb.close()
            return pd.DataFrame()

        # Find header rows — look for '学校' and '周'
        hdr_row = 0
        for i, row in enumerate(all_rows[:5]):
            row_str = ' '.join(str(v) for v in row if v)
            if '学校' in row_str:
                hdr_row = i
                break

        hdr = all_rows[hdr_row] if hdr_row < len(all_rows) else all_rows[0]
        hdr2 = all_rows[hdr_row+1] if hdr_row+1 < len(all_rows) else [None]*len(hdr)

        # Map columns
        col_map = {}
        for ci, (h, h2) in enumerate(zip(hdr, hdr2)):
            combined = str(h or '') + str(h2 or '')
            if '路线' in combined or '分区' in combined: col_map['route'] = ci
            elif '学校' in combined and 'route' not in col_map.get('school',''): col_map['school'] = ci
            elif '班级' in combined: col_map['class'] = ci
            elif '份数' in combined and '订' in combined: col_map['subs'] = ci
            elif '套餐' in combined: col_map['pkg'] = ci
            elif '规格' in combined: col_map['spec'] = ci
            elif '合计' in combined and '周' in combined: col_map['wk_total'] = ci
            elif '纯' in combined and '周' in combined: col_map['wk_pure'] = ci
            elif '乳' in combined and '周' in combined and '乳酸' in combined: col_map['wk_lactic'] = ci
            elif '甜' in combined: col_map['wk_sweet'] = ci
            elif '草莓' in combined or '草' in combined: col_map['wk_straw'] = ci
            elif '麦' in combined: col_map['wk_wheat'] = ci

        # Also do a broader scan: week supply columns
        # Column group: 周供应量 area has 纯/乳/甜/草/麦/合计
        # Scan row 1-2 for '周供应量' area
        wk_area_start = None
        for ci, v in enumerate(hdr):
            if '周供应' in str(v or ''):
                wk_area_start = ci
                break

        if wk_area_start and 'wk_total' not in col_map:
            # sub-headers in row 2 within this area
            flavor_map = {'纯': 'wk_pure', '乳': 'wk_lactic', '甜': 'wk_sweet',
                          '草': 'wk_straw', '麦': 'wk_wheat', '合': 'wk_total'}
            for ci in range(wk_area_start, min(wk_area_start+10, len(hdr2))):
                if ci < len(hdr2) and hdr2[ci]:
                    for k, v in flavor_map.items():
                        if k in str(hdr2[ci]):
                            col_map[v] = ci
                            break

        data_start = hdr_row + 2
        records = []
        for row in all_rows[data_start:]:
            if not row or len(row) < 3: continue
            school_v = row[col_map.get('school', 2)] if len(row) > col_map.get('school', 2) else None
            if not school_v: continue
            school_str = str(school_v).strip()
            if not school_str or school_str == 'None': continue
            if not re.search(r'[一-鿿]', school_str): continue

            def gv(key, default=0):
                ci = col_map.get(key)
                if ci is None or ci >= len(row): return default
                v = row[ci]
                if v is None: return default
                try: return int(float(str(v))) if key not in ('route','class','pkg','spec') else str(v).strip()
                except: return default

            records.append({
                '路线':   gv('route', ''),
                '学校':   school_str,
                '班级':   gv('class', ''),
                '套餐':   gv('pkg', ''),
                '规格':   gv('spec', ''),
                '征订份数': gv('subs'),
                '周_纯奶': gv('wk_pure'),
                '周_乳酸': gv('wk_lactic'),
                '周_甜奶': gv('wk_sweet'),
                '周_草莓': gv('wk_straw'),
                '周_麦香': gv('wk_wheat'),
                '周合计':  gv('wk_total'),
            })

        wb.close()
        df = pd.DataFrame(records)
        df = df[df['征订份数'] > 0]
        return df
    except Exception as e:
        print(f"  [ERR] 学生奶解析: {e}")
        traceback.print_exc()
        return pd.DataFrame()

student_milk_df = parse_student_milk(master_path)
print(f"[常温奶] 解析学生奶记录: {len(student_milk_df)} 条")

# ═══════════════════════════════════════════════════════════════════════════════
# PART 4 — Student milk detail from 英德中学，广德配货表.xlsx
# ═══════════════════════════════════════════════════════════════════════════════

def parse_delivery_detail(path):
    """Parse detailed student milk order from delivery app export."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        records = []
        for sname in wb.sheetnames:
            ws = wb[sname]
            all_rows = list(ws.iter_rows(values_only=True))
            # Find header row with '班级'
            hdr_idx = None
            for i, row in enumerate(all_rows[:8]):
                if any('班级' in str(v) for v in row if v):
                    hdr_idx = i; break
            if hdr_idx is None: continue
            hdr = all_rows[hdr_idx]
            col_class = next((ci for ci, v in enumerate(hdr) if '班级' in str(v or '')), 0)
            col_name  = next((ci for ci, v in enumerate(hdr) if '姓名' in str(v or '')), 1)
            col_pkg   = next((ci for ci, v in enumerate(hdr) if '套餐' in str(v or '')), 3)
            col_subs  = next((ci for ci, v in enumerate(hdr) if '份数' in str(v or '') and '合计' not in str(v or '')), 6)
            col_pure  = next((ci for ci, v in enumerate(hdr) if '纯奶' in str(v or '') or ('纯' in str(v or '') and '周' not in str(v or ''))), None)
            col_lactic= next((ci for ci, v in enumerate(hdr) if '乳酸' in str(v or '')), None)
            col_sweet = next((ci for ci, v in enumerate(hdr) if '甜奶' in str(v or '') or '甜' == str(v or '').strip()), None)
            col_straw = next((ci for ci, v in enumerate(hdr) if '草莓' in str(v or '')), None)
            col_wheat = next((ci for ci, v in enumerate(hdr) if '麦香' in str(v or '')), None)
            col_wk    = next((ci for ci, v in enumerate(hdr) if '合计' in str(v or '') and '周' in str(v or '')), None)

            def safe_int(row, ci):
                if ci is None or ci >= len(row): return 0
                try: return int(float(str(row[ci]))) if row[ci] is not None else 0
                except: return 0

            for row in all_rows[hdr_idx+1:]:
                if not row or len(row) < 3: continue
                cls = str(row[col_class] or '').strip() if col_class < len(row) else ''
                name = str(row[col_name] or '').strip() if col_name < len(row) else ''
                if not cls and not name: continue
                if not re.search(r'[一-鿿]', cls + name): continue
                records.append({
                    '学校': sname,
                    '班级': cls,
                    '姓名': name,
                    '套餐': str(row[col_pkg] or '').strip() if col_pkg < len(row) else '',
                    '份数': safe_int(row, col_subs),
                    '周_纯奶': safe_int(row, col_pure),
                    '周_乳酸': safe_int(row, col_lactic),
                    '周_甜奶': safe_int(row, col_sweet),
                    '周_草莓': safe_int(row, col_straw),
                    '周_麦香': safe_int(row, col_wheat),
                    '周合计':  safe_int(row, col_wk),
                })
        wb.close()
        return pd.DataFrame(records)
    except Exception as e:
        print(f"  [ERR] 配货表: {e}")
        return pd.DataFrame()

detail_path = os.path.join(BASE_DIR, "英德中学，广德配货表.xlsx")
detail_df = parse_delivery_detail(detail_path)
print(f"[常温奶详情] 解析记录: {len(detail_df)} 个学生")

# ═══════════════════════════════════════════════════════════════════════════════
# PART 5 — Gift redelivery list
# ═══════════════════════════════════════════════════════════════════════════════

def parse_gift_list(path):
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        # Find header
        hdr_idx = next((i for i, r in enumerate(rows) if any('学校' in str(v) for v in r if v)), 2)
        records = []
        for row in rows[hdr_idx+1:]:
            if not row or not any(row[:4]): continue
            records.append({
                '学校': str(row[0] or '').strip(),
                '职工姓名': str(row[1] or '').strip(),
                '礼品份数': row[2] if row[2] else 1,
            })
        return [r for r in records if r['学校'] or r['职工姓名']]
    except Exception as e:
        print(f"  [ERR] 礼品: {e}")
        return []

gift_records = parse_gift_list(os.path.join(BASE_DIR, "礼品补发名单.xlsx"))
print(f"[礼品] 补发名单: {len(gift_records)} 条")

# ═══════════════════════════════════════════════════════════════════════════════
# PART 6 — Merge delivery quantities back to school_df
# ═══════════════════════════════════════════════════════════════════════════════

# Try to match school_df schools with delivery_df schools
# For schools not in order files, use calculated values from subscription counts

def fuzzy_match(name, candidates):
    """Simple fuzzy match: find best match in candidates."""
    name = str(name).strip()
    for c in candidates:
        if name == c: return c
    for c in candidates:
        if name in c or c in name: return c
    # Try removing common suffixes
    for suffix in ['（内宿）','（外宿）','高中部','初中部','中学','小学','幼儿园']:
        stripped = name.replace(suffix, '')
        for c in candidates:
            if stripped in c or c in stripped: return c
    return None

order_schools = delivery_df['学校'].tolist() if '学校' in delivery_df.columns else []

# Add matched order quantities to school_df
for col in ['周一_鲜','周一_酸','周一_合计','周四_鲜','周四_酸','周四_合计','周合计_订单']:
    school_df[col] = 0

for idx, row in school_df.iterrows():
    matched = fuzzy_match(row['学校'], order_schools)
    if matched:
        orow = delivery_df[delivery_df['学校'] == matched].iloc[0]
        school_df.at[idx, '周一_鲜']      = orow.get('周一_鲜', 0)
        school_df.at[idx, '周一_酸']      = orow.get('周一_酸', 0)
        school_df.at[idx, '周一_合计']    = orow.get('周一_合计', 0)
        school_df.at[idx, '周四_鲜']      = orow.get('周四_鲜', 0)
        school_df.at[idx, '周四_酸']      = orow.get('周四_酸', 0)
        school_df.at[idx, '周四_合计']    = orow.get('周四_合计', 0)
        school_df.at[idx, '周合计_订单']  = orow.get('周合计', 0)
    else:
        # Use calculated values
        school_df.at[idx, '周一_鲜']   = row.get('mon_fresh', 0)
        school_df.at[idx, '周一_酸']   = row.get('mon_lactic', 0)
        school_df.at[idx, '周一_合计'] = row.get('mon_fresh', 0) + row.get('mon_lactic', 0)
        school_df.at[idx, '周四_鲜']   = row.get('thu_fresh', 0)
        school_df.at[idx, '周四_酸']   = row.get('thu_lactic', 0)
        school_df.at[idx, '周四_合计'] = row.get('thu_fresh', 0) + row.get('thu_lactic', 0)
        school_df.at[idx, '周合计_订单'] = (row.get('mon_fresh',0)+row.get('mon_lactic',0)+
                                             row.get('thu_fresh',0)+row.get('thu_lactic',0))

# ═══════════════════════════════════════════════════════════════════════════════
# BUILD EXCEL WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════

print("\n[Excel] 开始生成工作簿...")

wb_out = Workbook()
wb_out.remove(wb_out.active)  # remove default sheet

TODAY = datetime.date.today()
SEMESTER_START = datetime.date(2026, 3, 9)  # First Monday delivery
SEMESTER_END   = datetime.date(2026, 6, 28) # Estimated end

def get_delivery_weeks():
    """Generate all Mon/Thu delivery pairs from today onward."""
    weeks = []
    # Start from semester start Monday
    d = SEMESTER_START
    while d.weekday() != 0: d += datetime.timedelta(1)  # find Monday
    while d <= SEMESTER_END:
        mon = d
        thu = d + datetime.timedelta(3)
        sun = d - datetime.timedelta(1)  # Sunday before Monday (replenishment)
        wed = d + datetime.timedelta(2)  # Wednesday (replenishment for Thursday)
        weeks.append({'week_mon': mon, 'mon': mon, 'thu': thu, 'sun': sun, 'wed': wed})
        d += datetime.timedelta(7)
    return weeks

delivery_weeks = get_delivery_weeks()

# ────────────────────────────────────────────────────────────────────────────
# SHEET 1: 鲜奶征订汇总（按路线/学校）
# ────────────────────────────────────────────────────────────────────────────
ws1 = wb_out.create_sheet("①鲜奶征订汇总")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A4"

COLS1 = 13
set_col_widths(ws1, [4, 6, 22, 8, 8, 8, 8, 9, 9, 9, 9, 10, 16])

# Title
ws1.merge_cells("A1:M1")
write(ws1, 1, 1, "2026春季 鲜奶征订汇总表（教职工）", fill="navy", font="title")
ws1.row_dimensions[1].height = 32

# Sub-info
ws1.merge_cells("A2:M2")
info_cell = ws1.cell(row=2, column=1,
    value=f"统计日期: {TODAY}  |  鲜奶每周配送2次: 周一(3天量) + 周四(4天量)  |  1份=100支(鲜奶只选一种) 或 50鲜+50酸(两种均选)")
info_cell.fill = FILL["lblue"]
info_cell.font = Font(size=9, color="1F4E79")
info_cell.alignment = ALIGN["l"]
ws1.row_dimensions[2].height = 16

# Header row 3
for c, v in enumerate(["路线","序号","学校名称","征订人数","仅鲜奶\n(份)","仅乳酸\n(份)",
                        "鲜+酸\n混合(份)","合计份数","周一(3天)\n鲜奶(盒)","周一(3天)\n乳酸(盒)",
                        "周四(4天)\n鲜奶(盒)","周四(4天)\n乳酸(盒)","每周合计\n(盒)"], 1):
    write(ws1, 3, c+1, v, fill="navy", font="hdr")
ws1.row_dimensions[3].height = 36

# Data rows
row = 4
grand_total = defaultdict(int)
for route_name in ROUTE_ORDER:
    rdf = school_df[school_df['路线'] == route_name].reset_index(drop=True)
    if rdf.empty: continue

    hfill, dfill = ROUTE_FILLS[route_name]

    # Route header
    ws1.merge_cells(f"A{row}:M{row}")
    route_totals = rdf[['people','total_subs','fresh_only','lactic_only','mixed',
                         '周一_鲜','周一_酸','周四_鲜','周四_酸','周合计_订单']].sum()
    label = (f"【{route_name}】  共 {len(rdf)} 所学校  |  "
             f"征订人数: {int(route_totals['people'])}人  |  "
             f"合计份数: {int(route_totals['total_subs'])}份  |  "
             f"每周合计: {int(route_totals['周合计_订单'])}盒")
    write(ws1, row, 1, label, fill=hfill, font="route", align="l")
    ws1.row_dimensions[row].height = 20
    row += 1

    for i, (_, r) in enumerate(rdf.iterrows()):
        fill_key = dfill if i % 2 == 0 else "white"
        write(ws1, row, 1, route_name, fill=fill_key, font="small")
        write(ws1, row, 2, i+1, fill=fill_key, font="normal")
        write(ws1, row, 3, r['学校'], fill=fill_key, font="normal", align="l")
        write(ws1, row, 4, int(r.get('people',0)) or '', fill=fill_key)
        write(ws1, row, 5, int(r.get('fresh_only',0)) or '', fill=fill_key)
        write(ws1, row, 6, int(r.get('lactic_only',0)) or '', fill=fill_key)
        write(ws1, row, 7, int(r.get('mixed',0)) or '', fill=fill_key)
        subs = int(r.get('total_subs',0))
        write(ws1, row, 8, subs or '', fill=fill_key, font="bold" if subs else "normal")
        write(ws1, row, 9,  int(r.get('周一_鲜',0)) or '', fill=fill_key)
        write(ws1, row, 10, int(r.get('周一_酸',0)) or '', fill=fill_key)
        write(ws1, row, 11, int(r.get('周四_鲜',0)) or '', fill=fill_key)
        write(ws1, row, 12, int(r.get('周四_酸',0)) or '', fill=fill_key)
        wk = int(r.get('周合计_订单',0))
        write(ws1, row, 13, wk or '', fill=fill_key, font="bold" if wk else "normal")
        ws1.row_dimensions[row].height = 18
        row += 1

        for k in ['people','total_subs','fresh_only','lactic_only','mixed',
                  '周一_鲜','周一_酸','周四_鲜','周四_酸','周合计_订单']:
            grand_total[k] += int(r.get(k, 0))

    # Route subtotal
    for c in range(1, COLS1+1): ws1.cell(row, c).border = BORDER
    ws1.merge_cells(f"A{row}:C{row}")
    write(ws1, row, 1, f"【{route_name}】小计", fill="gray", font="bold", align="r")
    write(ws1, row, 4, int(route_totals['people']),       fill="gray", font="bold")
    write(ws1, row, 5, int(route_totals['fresh_only']),   fill="gray", font="bold")
    write(ws1, row, 6, int(route_totals['lactic_only']),  fill="gray", font="bold")
    write(ws1, row, 7, int(route_totals['mixed']),        fill="gray", font="bold")
    write(ws1, row, 8, int(route_totals['total_subs']),   fill="gray", font="bold")
    write(ws1, row, 9,  int(route_totals['周一_鲜']),     fill="gray", font="bold")
    write(ws1, row, 10, int(route_totals['周一_酸']),     fill="gray", font="bold")
    write(ws1, row, 11, int(route_totals['周四_鲜']),     fill="gray", font="bold")
    write(ws1, row, 12, int(route_totals['周四_酸']),     fill="gray", font="bold")
    write(ws1, row, 13, int(route_totals['周合计_订单']), fill="gray", font="bold")
    ws1.row_dimensions[row].height = 20
    row += 1

# Grand total
ws1.merge_cells(f"A{row}:C{row}")
write(ws1, row, 1, "全部路线合计", fill="red", font="hdr", align="r")
write(ws1, row, 4, grand_total['people'],     fill="red", font="hdr")
write(ws1, row, 5, grand_total['fresh_only'], fill="red", font="hdr")
write(ws1, row, 6, grand_total['lactic_only'],fill="red", font="hdr")
write(ws1, row, 7, grand_total['mixed'],      fill="red", font="hdr")
write(ws1, row, 8, grand_total['total_subs'], fill="red", font="hdr")
write(ws1, row, 9,  grand_total['周一_鲜'],   fill="red", font="hdr")
write(ws1, row, 10, grand_total['周一_酸'],   fill="red", font="hdr")
write(ws1, row, 11, grand_total['周四_鲜'],   fill="red", font="hdr")
write(ws1, row, 12, grand_total['周四_酸'],   fill="red", font="hdr")
write(ws1, row, 13, grand_total['周合计_订单'],fill="red", font="hdr")
ws1.row_dimensions[row].height = 22

print("  [✓] Sheet 1 完成")

# ────────────────────────────────────────────────────────────────────────────
# SHEET 2: 鲜奶周补货计划
# ────────────────────────────────────────────────────────────────────────────
ws2 = wb_out.create_sheet("②鲜奶补货计划")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A5"

# Weekly totals per route
route_weekly = {}
for route_name in ROUTE_ORDER:
    rdf = school_df[school_df['路线'] == route_name]
    route_weekly[route_name] = {
        'mon_fresh':  int(rdf['周一_鲜'].sum()),
        'mon_lactic': int(rdf['周一_酸'].sum()),
        'mon_total':  int(rdf['周一_合计'].sum()),
        'thu_fresh':  int(rdf['周四_鲜'].sum()),
        'thu_lactic': int(rdf['周四_酸'].sum()),
        'thu_total':  int(rdf['周四_合计'].sum()),
        'week_total': int(rdf['周合计_订单'].sum()),
    }

total_mon_fresh  = sum(v['mon_fresh'] for v in route_weekly.values())
total_mon_lactic = sum(v['mon_lactic'] for v in route_weekly.values())
total_mon        = sum(v['mon_total'] for v in route_weekly.values())
total_thu_fresh  = sum(v['thu_fresh'] for v in route_weekly.values())
total_thu_lactic = sum(v['thu_lactic'] for v in route_weekly.values())
total_thu        = sum(v['thu_total'] for v in route_weekly.values())
total_week       = total_mon + total_thu

# Use order file totals if available and higher
if not delivery_df.empty:
    df_mon = int(delivery_df['周一_合计'].sum()) if '周一_合计' in delivery_df.columns else 0
    df_thu = int(delivery_df['周四_合计'].sum()) if '周四_合计' in delivery_df.columns else 0
    df_mon_f = int(delivery_df['周一_鲜'].sum()) if '周一_鲜' in delivery_df.columns else 0
    df_mon_a = int(delivery_df['周一_酸'].sum()) if '周一_酸' in delivery_df.columns else 0
    df_thu_f = int(delivery_df['周四_鲜'].sum()) if '周四_鲜' in delivery_df.columns else 0
    df_thu_a = int(delivery_df['周四_酸'].sum()) if '周四_酸' in delivery_df.columns else 0
    if df_mon + df_thu > total_week:
        total_mon_fresh = df_mon_f; total_mon_lactic = df_mon_a; total_mon = df_mon
        total_thu_fresh = df_thu_f; total_thu_lactic = df_thu_a; total_thu = df_thu
        total_week = df_mon + df_thu

set_col_widths(ws2, [6, 14, 14, 14, 14, 12, 12, 12, 12, 14, 14])
COLS2 = 11

ws2.merge_cells("A1:K1")
write(ws2, 1, 1, "2026春季 鲜奶补货计划  |  到货节奏: 周日备货→周一配送(3天量)  /  周三备货→周四配送(4天量)",
      fill="navy", font="title")
ws2.row_dimensions[1].height = 30

# Section A: Weekly constant quantities (per route)
ws2.merge_cells("A2:K2")
write(ws2, 2, 1, "▌ A. 每周固定备货量（按路线分解）—— 订购量以本期实际征订份数为准，如有退订请同步调整",
      fill="blue", font="hdr", align="l")
ws2.row_dimensions[2].height = 20

# Header
hdrs2 = ["路线","周日备货\n鲜奶(盒)","周日备货\n乳酸(盒)","周日备货\n合计(盒)",
         "周三备货\n鲜奶(盒)","周三备货\n乳酸(盒)","周三备货\n合计(盒)",
         "每周配送\n合计(盒)","换算\n箱数(24盒)","月参考量\n(×4.3周)","备注"]
for c, v in enumerate(hdrs2, 1):
    write(ws2, 3, c, v, fill="navy", font="hdr")
ws2.row_dimensions[3].height = 36

rw = 4
for route_name in ROUTE_ORDER:
    rw_data = route_weekly[route_name]
    hfill = ROUTE_FILLS[route_name][0]
    mon_t = rw_data['mon_total']
    thu_t = rw_data['thu_total']
    wk_t  = rw_data['week_total']
    boxes = round(wk_t / 24, 1)
    month = round(wk_t * 4.3)
    write(ws2, rw, 1,  route_name, fill=hfill, font="hdr")
    write(ws2, rw, 2,  rw_data['mon_fresh'],  fill=hfill, font="hdr")
    write(ws2, rw, 3,  rw_data['mon_lactic'], fill=hfill, font="hdr")
    write(ws2, rw, 4,  mon_t,                 fill=hfill, font="hdr")
    write(ws2, rw, 5,  rw_data['thu_fresh'],  fill=hfill, font="hdr")
    write(ws2, rw, 6,  rw_data['thu_lactic'], fill=hfill, font="hdr")
    write(ws2, rw, 7,  thu_t,                 fill=hfill, font="hdr")
    write(ws2, rw, 8,  wk_t,                  fill=hfill, font="hdr")
    write(ws2, rw, 9,  boxes,                 fill=hfill, font="hdr")
    write(ws2, rw, 10, month,                 fill=hfill, font="hdr")
    write(ws2, rw, 11, '',                    fill=hfill, font="hdr")
    ws2.row_dimensions[rw].height = 22
    rw += 1

# Grand total row
write(ws2, rw, 1, "全部合计", fill="red", font="hdr")
write(ws2, rw, 2, total_mon_fresh,  fill="red", font="hdr")
write(ws2, rw, 3, total_mon_lactic, fill="red", font="hdr")
write(ws2, rw, 4, total_mon,        fill="red", font="hdr")
write(ws2, rw, 5, total_thu_fresh,  fill="red", font="hdr")
write(ws2, rw, 6, total_thu_lactic, fill="red", font="hdr")
write(ws2, rw, 7, total_thu,        fill="red", font="hdr")
write(ws2, rw, 8, total_week,       fill="red", font="hdr")
write(ws2, rw, 9, round(total_week/24,1), fill="red", font="hdr")
write(ws2, rw, 10, round(total_week*4.3), fill="red", font="hdr")
write(ws2, rw, 11, '',              fill="red", font="hdr")
ws2.row_dimensions[rw].height = 22
rw += 2

# Section B: Week-by-week schedule
ws2.merge_cells(f"A{rw}:K{rw}")
write(ws2, rw, 1,
      "▌ B. 逐周配送日历（2026-03-09 至 2026-06-28）—— 灰色=已过 / 黄色=本周 / 白色=待执行",
      fill="blue", font="hdr", align="l")
ws2.row_dimensions[rw].height = 20
rw += 1

cal_hdrs = ["周次","周日(备货)","周一(配送)","周一_鲜(盒)","周一_酸(盒)","周一合计",
            "周三(备货)","周四(配送)","周四_鲜(盒)","周四_酸(盒)","周四合计"]
for c, v in enumerate(cal_hdrs, 1):
    write(ws2, rw, c, v, fill="navy", font="hdr")
ws2.row_dimensions[rw].height = 30
rw += 1

for wi, wk in enumerate(delivery_weeks, 1):
    is_past    = wk['mon'] < TODAY
    is_current = wk['mon'] <= TODAY <= wk['thu'] + datetime.timedelta(1)
    fill_key = "lgray" if is_past else ("yellow" if is_current else "white")
    font_key = "small" if is_past else "normal"

    write(ws2, rw, 1, f"第{wi}周",          fill=fill_key, font=font_key)
    write(ws2, rw, 2, wk['sun'].strftime('%m/%d'), fill=fill_key, font=font_key)
    write(ws2, rw, 3, wk['mon'].strftime('%m/%d(%a)').replace('Mon','周一'), fill=fill_key, font=font_key)
    write(ws2, rw, 4, total_mon_fresh  if not is_past else '', fill=fill_key, font=font_key)
    write(ws2, rw, 5, total_mon_lactic if not is_past else '', fill=fill_key, font=font_key)
    write(ws2, rw, 6, total_mon        if not is_past else '', fill=fill_key, font="bold" if not is_past else "small")
    write(ws2, rw, 7, wk['wed'].strftime('%m/%d'), fill=fill_key, font=font_key)
    write(ws2, rw, 8, wk['thu'].strftime('%m/%d(%a)').replace('Thu','周四'), fill=fill_key, font=font_key)
    write(ws2, rw, 9,  total_thu_fresh  if not is_past else '', fill=fill_key, font=font_key)
    write(ws2, rw, 10, total_thu_lactic if not is_past else '', fill=fill_key, font=font_key)
    write(ws2, rw, 11, total_thu        if not is_past else '', fill=fill_key, font="bold" if not is_past else "small")
    ws2.row_dimensions[rw].height = 16
    rw += 1

print("  [✓] Sheet 2 完成")

# ────────────────────────────────────────────────────────────────────────────
# SHEET 3: 鲜奶各路线配送明细
# ────────────────────────────────────────────────────────────────────────────
ws3 = wb_out.create_sheet("③鲜奶配送明细")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A4"

set_col_widths(ws3, [6, 4, 26, 8, 8, 10, 8, 8, 10, 10, 14])
COLS3 = 11

ws3.merge_cells("A1:K1")
write(ws3, 1, 1, "2026春季 鲜奶配送明细（各学校 周一+周四 逐日配送量）",
      fill="navy", font="title")
ws3.row_dimensions[1].height = 30

ws3.merge_cells("A2:K2")
write(ws3, 2, 1,
      "说明: 每学校每周配送2次 | 周一送3天量(按每份3盒计) | 周四送4天量(按每份4盒计) | 订单量优先以实际系统导入数据为准",
      fill="lblue", font="small", align="l")
ws3.row_dimensions[2].height = 14

for c, v in enumerate(["路线","序号","学校名称",
                        "周一_鲜(盒)","周一_酸(盒)","周一合计(盒)",
                        "周四_鲜(盒)","周四_酸(盒)","周四合计(盒)",
                        "每周合计(盒)","数据来源"], 1):
    write(ws3, 3, c, v, fill="navy", font="hdr")
ws3.row_dimensions[3].height = 30

row = 4
for route_name in ROUTE_ORDER:
    rdf = school_df[school_df['路线'] == route_name].reset_index(drop=True)
    if rdf.empty: continue
    hfill, dfill = ROUTE_FILLS[route_name]

    # Route banner
    ws3.merge_cells(f"A{row}:K{row}")
    rt = rdf[['周一_鲜','周一_酸','周一_合计','周四_鲜','周四_酸','周四_合计','周合计_订单']].sum()
    write(ws3, row, 1,
          f"【{route_name}】{len(rdf)}所学校 | 周一合计: {int(rt['周一_合计'])}盒 | 周四合计: {int(rt['周四_合计'])}盒 | 每周: {int(rt['周合计_订单'])}盒",
          fill=hfill, font="route", align="l")
    ws3.row_dimensions[row].height = 20
    row += 1

    for i, (_, r) in enumerate(rdf.iterrows()):
        fill_key = dfill if i % 2 == 0 else "white"
        mon_f  = int(r['周一_鲜'])
        mon_a  = int(r['周一_酸'])
        mon_t  = int(r['周一_合计'])
        thu_f  = int(r['周四_鲜'])
        thu_a  = int(r['周四_酸'])
        thu_t  = int(r['周四_合计'])
        wk_t   = int(r['周合计_订单'])
        has_order = mon_t > 0 or thu_t > 0

        # Data source flag
        matched = fuzzy_match(r['学校'], order_schools)
        src = "订单导入" if matched else ("计算估算" if has_order else "待核实")

        write(ws3, row, 1, route_name, fill=fill_key, font="small")
        write(ws3, row, 2, i+1, fill=fill_key)
        write(ws3, row, 3, r['学校'], fill=fill_key, font="normal", align="l")
        write(ws3, row, 4, mon_f or '', fill=fill_key)
        write(ws3, row, 5, mon_a or '', fill=fill_key)
        write(ws3, row, 6, mon_t or '', fill=fill_key, font="bold" if mon_t else "normal")
        write(ws3, row, 7, thu_f or '', fill=fill_key)
        write(ws3, row, 8, thu_a or '', fill=fill_key)
        write(ws3, row, 9, thu_t or '', fill=fill_key, font="bold" if thu_t else "normal")
        write(ws3, row, 10, wk_t or '', fill=fill_key, font="bold" if wk_t else "normal")
        write(ws3, row, 11, src, fill="lgreen" if src=="订单导入" else ("yellow" if "估算" in src else "lred"),
              font="small")
        ws3.row_dimensions[row].height = 18
        row += 1

    # Subtotal
    ws3.merge_cells(f"A{row}:C{row}")
    write(ws3, row, 1, f"【{route_name}】小计", fill="gray", font="bold", align="r")
    write(ws3, row, 4, int(rt['周一_鲜']),   fill="gray", font="bold")
    write(ws3, row, 5, int(rt['周一_酸']),   fill="gray", font="bold")
    write(ws3, row, 6, int(rt['周一_合计']), fill="gray", font="bold")
    write(ws3, row, 7, int(rt['周四_鲜']),   fill="gray", font="bold")
    write(ws3, row, 8, int(rt['周四_酸']),   fill="gray", font="bold")
    write(ws3, row, 9, int(rt['周四_合计']), fill="gray", font="bold")
    write(ws3, row, 10, int(rt['周合计_订单']), fill="gray", font="bold")
    write(ws3, row, 11, '', fill="gray")
    ws3.row_dimensions[row].height = 20
    row += 1

# Grand total
ws3.merge_cells(f"A{row}:C{row}")
write(ws3, row, 1, "全部路线合计", fill="red", font="hdr", align="r")
write(ws3, row, 4, total_mon_fresh,  fill="red", font="hdr")
write(ws3, row, 5, total_mon_lactic, fill="red", font="hdr")
write(ws3, row, 6, total_mon,        fill="red", font="hdr")
write(ws3, row, 7, total_thu_fresh,  fill="red", font="hdr")
write(ws3, row, 8, total_thu_lactic, fill="red", font="hdr")
write(ws3, row, 9, total_thu,        fill="red", font="hdr")
write(ws3, row, 10, total_week,      fill="red", font="hdr")
write(ws3, row, 11, '',              fill="red", font="hdr")
ws3.row_dimensions[row].height = 22

print("  [✓] Sheet 3 完成")

# ────────────────────────────────────────────────────────────────────────────
# SHEET 4: 常温奶征订统计
# ────────────────────────────────────────────────────────────────────────────
ws4 = wb_out.create_sheet("④常温奶征订统计")
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = "A4"

set_col_widths(ws4, [8, 4, 26, 16, 6, 8, 8, 8, 8, 8, 10, 12])
COLS4 = 12

ws4.merge_cells("A1:L1")
write(ws4, 1, 1, "2026春季 常温学生奶征订统计（家长订阅套餐 100支/套）",
      fill="green", font="title")
ws4.row_dimensions[1].height = 30

ws4.merge_cells("A2:L2")
write(ws4, 2, 1,
      "数据来源: 各学校明细配送表(学生奶) | 100支/套=约14周配送(7支/周) | 每周或每半月配送至学校奶室，由班主任分发",
      fill="lgreen", font="small", align="l")
ws4.row_dimensions[2].height = 14

for c, v in enumerate(["路线","序号","学校名称","班级/套餐","征订\n份数",
                        "周_纯奶\n(支)","周_乳酸\n(支)","周_甜奶\n(支)","周_草莓\n(支)","周_麦香\n(支)",
                        "周合计\n(支)","备注"], 1):
    write(ws4, 3, c, v, fill="green", font="hdr")
ws4.row_dimensions[3].height = 36

STUDENT_FILL = ("green", "lgreen")

if student_milk_df.empty:
    ws4.merge_cells(f"A4:L4")
    write(ws4, 4, 1, "⚠️ 未能从主文件解析到常温奶数据，请检查「2026年春季各学校明细配送表.xlsx」中的「学生奶」工作表",
          fill="lred", font="bold")
    ws4.row_dimensions[4].height = 24
    sm_total_subs = 0; sm_total_wk = 0
    row4 = 5
else:
    row4 = 4
    route_grp = student_milk_df.groupby('路线') if '路线' in student_milk_df.columns else None

    # Group by route if available, otherwise just list all
    if route_grp and student_milk_df['路线'].notna().any():
        all_routes_sm = student_milk_df['路线'].dropna().unique().tolist()
    else:
        all_routes_sm = ['（未分路线）']

    sm_grand_subs = 0; sm_grand_wk = 0
    seq = 0

    # Sort by route label
    sm_df_sorted = student_milk_df.copy()
    route_idx_map = {r: i for i, r in enumerate(ROUTE_ORDER)}
    if '路线' in sm_df_sorted.columns:
        sm_df_sorted['_route_order'] = sm_df_sorted['路线'].map(
            lambda x: next((i for r, i in route_idx_map.items() if r in str(x)), 99))
        sm_df_sorted = sm_df_sorted.sort_values(['_route_order','学校']).drop(columns='_route_order')

    cur_route = None
    for _, r in sm_df_sorted.iterrows():
        rname = str(r.get('路线', '')).strip() or '（未分路线）'
        if rname != cur_route:
            # Route header
            rt_sub = sm_df_sorted[sm_df_sorted['路线'].astype(str).str.strip() == rname] if '路线' in sm_df_sorted.columns else sm_df_sorted
            rt_subs = int(rt_sub['征订份数'].sum())
            rt_wk   = int(rt_sub['周合计'].sum())
            ws4.merge_cells(f"A{row4}:L{row4}")
            hfill = next((ROUTE_FILLS[k][0] for k in ROUTE_FILLS if k in rname), "blue")
            write(ws4, row4, 1,
                  f"【{rname}】 征订: {rt_subs}份 | 每周: {rt_wk}支",
                  fill=hfill, font="route", align="l")
            ws4.row_dimensions[row4].height = 20
            row4 += 1
            cur_route = rname
            seq = 0

        seq += 1
        alt = seq % 2 == 0
        dfill = "lgreen" if alt else "white"

        cls_val = str(r.get('班级', '')).strip() or str(r.get('套餐', '')).strip()[:30] or ''
        subs = int(r.get('征订份数', 0))
        wk_t = int(r.get('周合计', 0))

        write(ws4, row4, 1, rname, fill=dfill, font="small")
        write(ws4, row4, 2, seq, fill=dfill)
        write(ws4, row4, 3, str(r['学校'])[:30], fill=dfill, font="normal", align="l")
        write(ws4, row4, 4, cls_val[:30], fill=dfill, font="small", align="l")
        write(ws4, row4, 5, subs or '', fill=dfill, font="bold" if subs else "normal")
        write(ws4, row4, 6, int(r.get('周_纯奶', 0)) or '', fill=dfill)
        write(ws4, row4, 7, int(r.get('周_乳酸', 0)) or '', fill=dfill)
        write(ws4, row4, 8, int(r.get('周_甜奶', 0)) or '', fill=dfill)
        write(ws4, row4, 9, int(r.get('周_草莓', 0)) or '', fill=dfill)
        write(ws4, row4, 10, int(r.get('周_麦香', 0)) or '', fill=dfill)
        write(ws4, row4, 11, wk_t or '', fill=dfill, font="bold" if wk_t else "normal")
        write(ws4, row4, 12, '', fill=dfill)
        ws4.row_dimensions[row4].height = 16
        row4 += 1
        sm_grand_subs += subs; sm_grand_wk += wk_t

    # Grand total
    ws4.merge_cells(f"A{row4}:D{row4}")
    write(ws4, row4, 1, "全部学校合计", fill="red", font="hdr", align="r")
    write(ws4, row4, 5, sm_grand_subs, fill="red", font="hdr")
    for c in range(6, 11): write(ws4, row4, c, '', fill="red", font="hdr")
    write(ws4, row4, 11, sm_grand_wk, fill="red", font="hdr")
    write(ws4, row4, 12, '', fill="red", font="hdr")
    ws4.row_dimensions[row4].height = 22

    sm_total_subs = sm_grand_subs; sm_total_wk = sm_grand_wk

print("  [✓] Sheet 4 完成")

# ────────────────────────────────────────────────────────────────────────────
# SHEET 5: 常温奶配送计划
# ────────────────────────────────────────────────────────────────────────────
ws5 = wb_out.create_sheet("⑤常温奶配送计划")
ws5.sheet_view.showGridLines = False
ws5.freeze_panes = "A4"

set_col_widths(ws5, [8, 4, 26, 6, 6, 6, 6, 6, 8, 10, 12, 12])
COLS5 = 12

ws5.merge_cells("A1:L1")
write(ws5, 1, 1, "2026春季 常温学生奶 配送计划（学校奶室分发）",
      fill="green", font="title")
ws5.row_dimensions[1].height = 30

ws5.merge_cells("A2:L2")
write(ws5, 2, 1,
      "配送方式: 送至学校奶室 / 各班领取 | 建议每周或每半月配送一次 | 每次配送量=周量×配送间隔周数",
      fill="lgreen", font="small", align="l")
ws5.row_dimensions[2].height = 14

# Two sub-sections: weekly and bi-weekly
# Section A: weekly delivery by school
ws5.merge_cells("A3:L3")
write(ws5, 3, 1, "▌ 每周配送方案（每周一与常规鲜奶同批配送至学校）",
      fill="blue", font="hdr", align="l")
ws5.row_dimensions[3].height = 20

for c, v in enumerate(["路线","序号","学校","征订份数","周_纯","周_乳","周_甜","周_草","周_麦","周合计(支)",
                        "建议配送频次","备注"], 1):
    write(ws5, 4, c, v, fill="navy", font="hdr")
ws5.row_dimensions[4].height = 30

row5 = 5
sm_route_totals = defaultdict(lambda: defaultdict(int))

sm_df_out = sm_df_sorted if not student_milk_df.empty else pd.DataFrame()
cur_route5 = None; seq5 = 0

for _, r in sm_df_out.iterrows() if not sm_df_out.empty else []:
    rname = str(r.get('路线', '')).strip() or '（未分路线）'
    if rname != cur_route5:
        if cur_route5 is not None:
            # Subtotal for previous route
            st = sm_route_totals[cur_route5]
            ws5.merge_cells(f"A{row5}:C{row5}")
            hfill = next((ROUTE_FILLS[k][0] for k in ROUTE_FILLS if k in cur_route5), "blue")
            write(ws5, row5, 1, f"【{cur_route5}】小计", fill="gray", font="bold", align="r")
            write(ws5, row5, 4, st['subs'], fill="gray", font="bold")
            for c, k in enumerate(['pure','lac','sw','st','wh','tot'], 5):
                write(ws5, row5, c, st[k] or '', fill="gray", font="bold")
            write(ws5, row5, 11, ''); write(ws5, row5, 12, '')
            ws5.row_dimensions[row5].height = 18; row5 += 1

        hfill = next((ROUTE_FILLS[k][0] for k in ROUTE_FILLS if k in rname), "blue")
        ws5.merge_cells(f"A{row5}:L{row5}")
        rt_sub5 = sm_df_out[sm_df_out['路线'].astype(str).str.strip() == rname] if '路线' in sm_df_out.columns else sm_df_out
        rt_subs5 = int(rt_sub5['征订份数'].sum()); rt_wk5 = int(rt_sub5['周合计'].sum())
        write(ws5, row5, 1, f"【{rname}】 征订: {rt_subs5}份 | 周合计: {rt_wk5}支",
              fill=hfill, font="route", align="l")
        ws5.row_dimensions[row5].height = 20; row5 += 1
        cur_route5 = rname; seq5 = 0

    seq5 += 1
    dfill5 = "lgreen" if seq5 % 2 == 0 else "white"
    subs5 = int(r.get('征订份数', 0))
    pure5  = int(r.get('周_纯奶', 0)); lac5 = int(r.get('周_乳酸', 0))
    sw5    = int(r.get('周_甜奶', 0)); st5  = int(r.get('周_草莓', 0))
    wh5    = int(r.get('周_麦香', 0)); tot5 = int(r.get('周合计', 0))

    sm_route_totals[rname]['subs'] += subs5
    sm_route_totals[rname]['pure'] += pure5; sm_route_totals[rname]['lac']  += lac5
    sm_route_totals[rname]['sw']   += sw5;   sm_route_totals[rname]['st']   += st5
    sm_route_totals[rname]['wh']   += wh5;   sm_route_totals[rname]['tot']  += tot5

    # frequency recommendation: if weekly total > 50, deliver weekly; else bi-weekly
    freq = "每周" if tot5 >= 50 else "每两周"

    write(ws5, row5, 1, rname, fill=dfill5, font="small")
    write(ws5, row5, 2, seq5, fill=dfill5)
    write(ws5, row5, 3, str(r['学校'])[:30], fill=dfill5, font="normal", align="l")
    write(ws5, row5, 4, subs5 or '', fill=dfill5, font="bold" if subs5 else "normal")
    write(ws5, row5, 5, pure5 or '', fill=dfill5)
    write(ws5, row5, 6, lac5  or '', fill=dfill5)
    write(ws5, row5, 7, sw5   or '', fill=dfill5)
    write(ws5, row5, 8, st5   or '', fill=dfill5)
    write(ws5, row5, 9, wh5   or '', fill=dfill5)
    write(ws5, row5, 10, tot5 or '', fill=dfill5, font="bold" if tot5 else "normal")
    write(ws5, row5, 11, freq, fill=dfill5, font="small")
    write(ws5, row5, 12, '', fill=dfill5)
    ws5.row_dimensions[row5].height = 16; row5 += 1

if cur_route5 and not sm_df_out.empty:
    st = sm_route_totals[cur_route5]
    ws5.merge_cells(f"A{row5}:C{row5}")
    write(ws5, row5, 1, f"【{cur_route5}】小计", fill="gray", font="bold", align="r")
    write(ws5, row5, 4, st['subs'], fill="gray", font="bold")
    for c, k in enumerate(['pure','lac','sw','st','wh','tot'], 5):
        write(ws5, row5, c, st[k] or '', fill="gray", font="bold")
    write(ws5, row5, 11, ''); write(ws5, row5, 12, '')
    ws5.row_dimensions[row5].height = 18; row5 += 1

# Grand total
ws5.merge_cells(f"A{row5}:C{row5}")
write(ws5, row5, 1, "全部合计", fill="red", font="hdr", align="r")
write(ws5, row5, 4, sm_total_subs, fill="red", font="hdr")
all_pure = sum(sm_route_totals[r]['pure'] for r in sm_route_totals)
all_lac  = sum(sm_route_totals[r]['lac']  for r in sm_route_totals)
all_sw   = sum(sm_route_totals[r]['sw']   for r in sm_route_totals)
all_st   = sum(sm_route_totals[r]['st']   for r in sm_route_totals)
all_wh   = sum(sm_route_totals[r]['wh']   for r in sm_route_totals)
for c, v in enumerate([all_pure, all_lac, all_sw, all_st, all_wh, sm_total_wk], 5):
    write(ws5, row5, c, v or '', fill="red", font="hdr")
write(ws5, row5, 11, ''); write(ws5, row5, 12, '')
ws5.row_dimensions[row5].height = 22; row5 += 2

# Section B: Bi-weekly delivery reference
ws5.merge_cells(f"A{row5}:L{row5}")
write(ws5, row5, 1, "▌ 每半月配送参考量（每次 = 周量 × 2 = 约两周用量，建议与月中/月末常规货配合）",
      fill="blue", font="hdr", align="l")
ws5.row_dimensions[row5].height = 20; row5 += 1

for c, v in enumerate(["路线","序号","学校","征订份数","半月_纯","半月_乳","半月_甜","半月_草","半月_麦",
                        "半月合计(支)","换算箱数\n(约24支/箱)","备注"], 1):
    write(ws5, row5, c, v, fill="navy", font="hdr")
ws5.row_dimensions[row5].height = 30; row5 += 1

seq5b = 0; cur5b = None
for _, r in sm_df_out.iterrows() if not sm_df_out.empty else []:
    rname = str(r.get('路线', '')).strip() or '（未分路线）'
    if rname != cur5b:
        if cur5b:
            pass  # no subtotal for this section to keep compact
        hfill = next((ROUTE_FILLS[k][0] for k in ROUTE_FILLS if k in rname), "blue")
        ws5.merge_cells(f"A{row5}:L{row5}")
        write(ws5, row5, 1, f"【{rname}】", fill=hfill, font="route", align="l")
        ws5.row_dimensions[row5].height = 18; row5 += 1
        cur5b = rname; seq5b = 0

    seq5b += 1
    dfill5b = "lgreen" if seq5b % 2 == 0 else "white"
    subs5b = int(r.get('征订份数', 0))
    tot5b  = int(r.get('周合计', 0)) * 2
    pure5b = int(r.get('周_纯奶', 0)) * 2; lac5b = int(r.get('周_乳酸', 0)) * 2
    sw5b   = int(r.get('周_甜奶', 0)) * 2; st5b  = int(r.get('周_草莓', 0)) * 2
    wh5b   = int(r.get('周_麦香', 0)) * 2
    boxes5b = round(tot5b / 24, 1) if tot5b else ''

    write(ws5, row5, 1, rname, fill=dfill5b, font="small")
    write(ws5, row5, 2, seq5b, fill=dfill5b)
    write(ws5, row5, 3, str(r['学校'])[:30], fill=dfill5b, font="normal", align="l")
    write(ws5, row5, 4, subs5b or '', fill=dfill5b)
    write(ws5, row5, 5, pure5b or '', fill=dfill5b)
    write(ws5, row5, 6, lac5b  or '', fill=dfill5b)
    write(ws5, row5, 7, sw5b   or '', fill=dfill5b)
    write(ws5, row5, 8, st5b   or '', fill=dfill5b)
    write(ws5, row5, 9, wh5b   or '', fill=dfill5b)
    write(ws5, row5, 10, tot5b or '', fill=dfill5b, font="bold" if tot5b else "normal")
    write(ws5, row5, 11, boxes5b, fill=dfill5b)
    write(ws5, row5, 12, '', fill=dfill5b)
    ws5.row_dimensions[row5].height = 15; row5 += 1

print("  [✓] Sheet 5 完成")

# ────────────────────────────────────────────────────────────────────────────
# SHEET 6: 常温奶班级明细 (from 英德中学，广德配货表)
# ────────────────────────────────────────────────────────────────────────────
ws6 = wb_out.create_sheet("⑥常温奶班级明细")
ws6.sheet_view.showGridLines = False
ws6.freeze_panes = "A4"

set_col_widths(ws6, [14, 16, 14, 24, 4, 6, 6, 6, 6, 6, 8, 10])
COLS6 = 12

ws6.merge_cells("A1:L1")
write(ws6, 1, 1, "常温奶班级明细（英德中学高中部/初中部 · 广德）— 来源: 订奶系统导出",
      fill="green", font="title")
ws6.row_dimensions[1].height = 30

if detail_df.empty:
    ws6.merge_cells("A2:L2")
    write(ws6, 2, 1, "⚠️ 未找到「英德中学，广德配货表.xlsx」或数据为空",
          fill="lred", font="bold")
else:
    for c, v in enumerate(["学校","班级","学生姓名","套餐","份数",
                            "纯奶(周)","乳酸(周)","甜奶(周)","草莓(周)","麦香(周)","周合计","备注"], 1):
        write(ws6, 3, c, v, fill="green", font="hdr")
    ws6.row_dimensions[3].height = 30

    row6 = 4
    cur_school6 = None; seq6 = 0
    for _, r in detail_df.iterrows():
        sn = str(r.get('学校', '')).strip()
        if sn != cur_school6:
            if cur_school6:
                # subtotal
                sub6 = detail_df[detail_df['学校'] == cur_school6]
                ws6.merge_cells(f"A{row6}:D{row6}")
                write(ws6, row6, 1, f"【{cur_school6}】小计: {len(sub6)}人", fill="gray", font="bold", align="r")
                write(ws6, row6, 5, int(sub6['份数'].sum()), fill="gray", font="bold")
                for c, col in enumerate(['周_纯奶','周_乳酸','周_甜奶','周_草莓','周_麦香','周合计'], 6):
                    write(ws6, row6, c, int(sub6[col].sum()) if col in sub6.columns else '', fill="gray", font="bold")
                write(ws6, row6, 12, '', fill="gray")
                ws6.row_dimensions[row6].height = 18; row6 += 1

            ws6.merge_cells(f"A{row6}:L{row6}")
            write(ws6, row6, 1, f"▌ {sn}", fill="blue", font="route", align="l")
            ws6.row_dimensions[row6].height = 20; row6 += 1
            cur_school6 = sn; seq6 = 0

        seq6 += 1
        dfill6 = "lgreen" if seq6 % 2 == 0 else "white"
        write(ws6, row6, 1,  sn[:20],  fill=dfill6, font="small")
        write(ws6, row6, 2,  str(r.get('班级',''))[:20], fill=dfill6, font="normal", align="l")
        write(ws6, row6, 3,  str(r.get('姓名',''))[:20], fill=dfill6, font="normal", align="l")
        write(ws6, row6, 4,  str(r.get('套餐',''))[:30], fill=dfill6, font="small", align="l")
        write(ws6, row6, 5,  int(r.get('份数',0)) or '', fill=dfill6)
        write(ws6, row6, 6,  int(r.get('周_纯奶',0)) or '', fill=dfill6)
        write(ws6, row6, 7,  int(r.get('周_乳酸',0)) or '', fill=dfill6)
        write(ws6, row6, 8,  int(r.get('周_甜奶',0)) or '', fill=dfill6)
        write(ws6, row6, 9,  int(r.get('周_草莓',0)) or '', fill=dfill6)
        write(ws6, row6, 10, int(r.get('周_麦香',0)) or '', fill=dfill6)
        write(ws6, row6, 11, int(r.get('周合计',0)) or '', fill=dfill6, font="bold" if r.get('周合计') else "normal")
        write(ws6, row6, 12, '', fill=dfill6)
        ws6.row_dimensions[row6].height = 16; row6 += 1

    # Final subtotal
    if cur_school6:
        sub6 = detail_df[detail_df['学校'] == cur_school6]
        ws6.merge_cells(f"A{row6}:D{row6}")
        write(ws6, row6, 1, f"【{cur_school6}】小计: {len(sub6)}人", fill="gray", font="bold", align="r")
        write(ws6, row6, 5, int(sub6['份数'].sum()), fill="gray", font="bold")
        for c, col in enumerate(['周_纯奶','周_乳酸','周_甜奶','周_草莓','周_麦香','周合计'], 6):
            write(ws6, row6, c, int(sub6[col].sum()) if col in sub6.columns else '', fill="gray", font="bold")
        write(ws6, row6, 12, '', fill="gray")
        ws6.row_dimensions[row6].height = 18; row6 += 1

    # Grand total
    ws6.merge_cells(f"A{row6}:D{row6}")
    write(ws6, row6, 1, f"全部合计: {len(detail_df)}人", fill="red", font="hdr", align="r")
    write(ws6, row6, 5, int(detail_df['份数'].sum()), fill="red", font="hdr")
    for c, col in enumerate(['周_纯奶','周_乳酸','周_甜奶','周_草莓','周_麦香','周合计'], 6):
        write(ws6, row6, c, int(detail_df[col].sum()) if col in detail_df.columns else '', fill="red", font="hdr")
    write(ws6, row6, 12, '', fill="red", font="hdr")
    ws6.row_dimensions[row6].height = 22

print("  [✓] Sheet 6 完成")

# ────────────────────────────────────────────────────────────────────────────
# SHEET 7: 礼品补发名单
# ────────────────────────────────────────────────────────────────────────────
ws7 = wb_out.create_sheet("⑦礼品补发名单")
ws7.sheet_view.showGridLines = False

set_col_widths(ws7, [4, 28, 16, 10, 20])

ws7.merge_cells("A1:E1")
write(ws7, 1, 1, "礼品补发名单（笔记本+笔）", fill="purple", font="title")
ws7.row_dimensions[1].height = 30

ws7.merge_cells("A2:E2")
write(ws7, 2, 1, f"共 {len(gift_records)} 人 | 礼品内容: 笔记本 + 笔 | 每份1套",
      fill="lgray", font="bold")
ws7.row_dimensions[2].height = 18

for c, v in enumerate(["序号","学校","职工姓名","礼品份数","备注"], 1):
    write(ws7, 3, c, v, fill="purple", font="hdr")
ws7.row_dimensions[3].height = 26

for i, rec in enumerate(gift_records, 1):
    dfill7 = "lgray" if i % 2 == 0 else "white"
    write(ws7, 3+i, 1, i, fill=dfill7)
    write(ws7, 3+i, 2, rec.get('学校',''), fill=dfill7, font="normal", align="l")
    write(ws7, 3+i, 3, rec.get('职工姓名',''), fill=dfill7, font="normal")
    write(ws7, 3+i, 4, rec.get('礼品份数', 1), fill=dfill7, font="bold")
    write(ws7, 3+i, 5, '', fill=dfill7)
    ws7.row_dimensions[3+i].height = 18

total_gifts = sum(int(r.get('礼品份数', 1)) for r in gift_records)
tr = 4 + len(gift_records)
ws7.merge_cells(f"A{tr}:C{tr}")
write(ws7, tr, 1, "合计", fill="purple", font="hdr", align="r")
write(ws7, tr, 4, total_gifts, fill="purple", font="hdr")
write(ws7, tr, 5, '', fill="purple", font="hdr")
ws7.row_dimensions[tr].height = 22

print("  [✓] Sheet 7 完成")

# ────────────────────────────────────────────────────────────────────────────
# SHEET 8: 说明 / 使用指南
# ────────────────────────────────────────────────────────────────────────────
ws8 = wb_out.create_sheet("说明")
ws8.sheet_view.showGridLines = False
set_col_widths(ws8, [5, 20, 60, 18])

ws8.merge_cells("A1:D1")
write(ws8, 1, 1, "2026春季学生奶业务管理总表 · 使用说明", fill="navy", font="title")
ws8.row_dimensions[1].height = 35

notes = [
    ("工作表", "工作表名称", "说明", "对应业务场景"),
    ("①", "鲜奶征订汇总", "汇总各路线各学校教职工鲜奶征订情况，含人数、份数、每周周一/周四配送量分解", "征订管理 / 业务员核对"),
    ("②", "鲜奶补货计划", "按路线统计每周日/周三需备货量，含逐周日历（灰=已过/黄=本周/白=待执行）", "备货采购 / 供应商下单"),
    ("③", "鲜奶配送明细", "各学校周一+周四配送盒数明细（鲜/酸分列），标注数据来源（订单导入/估算）", "配送司机 / 路线分拣"),
    ("④", "常温奶征订统计", "各学校班级常温学生奶征订份数、各口味周供量，来源: 各学校明细配送表学生奶Sheet", "奶室管理 / 班主任分发"),
    ("⑤", "常温奶配送计划", "每周配送量+每半月配送换算，含建议配送频次", "送货计划 / 装车分拣"),
    ("⑥", "常温奶班级明细", "英德中学(高/初中部)及广德学校 学生级别订奶明细，精确到人", "班级分发 / 对账"),
    ("⑦", "礼品补发名单", "待补发笔记本+笔礼品的教职工名单", "礼品配送"),
    ("", "", "", ""),
    ("说明", "鲜奶配送节奏", "每周2次: 周日备货→周一配送(3天量，每份3盒) / 周三备货→周四配送(4天量，每份4盒)", ""),
    ("", "1份含义", "混合订购(鲜+酸均选): 1份=50盒鲜奶+50盒乳酸 | 单选鲜奶: 1份=100盒鲜奶 | 单选乳酸: 1份=100盒乳酸", ""),
    ("", "常温奶套餐", "100支/套, 家长线上订购, 约14周配送(7支/周), 送至学校奶室由班主任/奶室负责人分发", ""),
    ("", "数据更新", "订单量以最新系统导出的3天/4天录入明细总表为准; 征订变动请在源文件更新后重新运行脚本", ""),
    ("", "生成日期", str(TODAY), ""),
]
for i, row_data in enumerate(notes, 2):
    fills_n = ["gray","navy","white","white"] if i == 2 else (["lgray","lblue","white","white"] if i % 2 == 0 else ["lgray","dblue","white","white"])
    fonts_n = ["hdr","hdr","hdr","hdr"] if i == 2 else ["bold","bold","normal","normal"]
    for c, (v, f, fn) in enumerate(zip(row_data, fills_n, fonts_n), 1):
        write(ws8, i, c, v, fill=f, font=fn, align="l" if c >= 3 else "c")
    ws8.row_dimensions[i].height = 20 if i > 2 else 24

print("  [✓] Sheet 8 完成")

# ────────────────────────────────────────────────────────────────────────────
# Save
# ────────────────────────────────────────────────────────────────────────────
out_path = os.path.join(BASE_DIR, "2026春季学生奶业务管理总表.xlsx")
wb_out.save(out_path)
print(f"\n[完成] 已保存: {out_path}")
print(f"  工作表: {wb_out.sheetnames}")
print(f"  鲜奶学校数: {len(school_df)}")
print(f"  常温奶记录: {len(student_milk_df)}")
print(f"  学生详情: {len(detail_df)}")
print(f"  礼品补发: {len(gift_records)}")
